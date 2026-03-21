"""NATO Defence Expenditure connector.

Fetches and parses the NATO Defence Expenditure annual report Excel file.
Provides defence spending data for all NATO member countries (2014-2025).

Data includes:
  - Defence expenditure in million USD (current prices)
  - Defence expenditure in million USD (constant 2021 prices)
  - Defence expenditure as share of GDP (%)
  - Annual real change in defence expenditure (%)

The Excel file contains 9 worksheets. This connector focuses on:
  - Table 2: Defence expenditure (million USD)
  - Table 3: Defence expenditure as share of GDP / annual real change

Reference: https://www.nato.int/cps/en/natohq/topics_49198.htm
"""

from __future__ import annotations

import io
import logging
import re
import time
from dataclasses import dataclass, field

import httpx
import openpyxl

logger = logging.getLogger(__name__)

NATO_EXCEL_URL = (
    "https://www.nato.int/content/dam/nato/webready/documents/finance/def-exp-2025-en.xlsx"
)

# Rows containing individual country data (1-indexed) in Table 2 and Table 3.
# Rows 5-35 in the first section, 41-71 in the second section.
# Rows 36-37 / 72-73 are aggregates (NATO Europe+Canada, NATO Total).
_COUNTRY_ROWS_SECTION1 = range(5, 36)   # rows 5..35 inclusive
_COUNTRY_ROWS_SECTION2 = range(41, 72)  # rows 41..71 inclusive
_AGGREGATE_ROWS_SECTION1 = range(36, 38)
_AGGREGATE_ROWS_SECTION2 = range(72, 74)

# Year columns: B(2)=2014, C(3)=2015, ..., M(13)=2025e
_YEAR_COL_START = 2   # column B
_YEAR_COL_END = 13    # column M
_BASE_YEAR = 2014


@dataclass
class NATOSpendingRecord:
    """A single country-year NATO defence spending record."""

    country: str
    year: int
    spending_current_usd_millions: float | None
    spending_constant_usd_millions: float | None
    pct_gdp: float | None
    annual_real_change_pct: float | None
    is_estimate: bool
    is_aggregate: bool = False


@dataclass
class _CacheEntry:
    """Internal cache entry with TTL."""

    records: list[NATOSpendingRecord]
    fetched_at: float


class NATOSpendingClient:
    """Client for NATO Defence Expenditure data.

    Downloads and parses the annual Excel report from NATO.
    Results are cached in memory since the file is updated only once per year.
    """

    def __init__(self, timeout: float = 60.0, cache_ttl: float = 86400.0):
        self.timeout = timeout
        self.cache_ttl = cache_ttl
        self._cache: _CacheEntry | None = None

    async def fetch_excel(self, url: str = NATO_EXCEL_URL) -> bytes:
        """Download the NATO defence expenditure Excel file."""
        async with httpx.AsyncClient(timeout=self.timeout) as client:
            logger.info("Downloading NATO defence expenditure Excel from %s", url)
            response = await client.get(url, follow_redirects=True)
            response.raise_for_status()
            logger.info(
                "Downloaded %d bytes from NATO", len(response.content)
            )
            return response.content

    def parse_excel(self, excel_bytes: bytes) -> list[NATOSpendingRecord]:
        """Parse the NATO Excel file into structured records.

        Extracts data from:
          - Table 2: Defence expenditure in million USD
            * Section 1 (rows 5-37): Current prices and exchange rates
            * Section 2 (rows 41-73): Constant 2021 prices and exchange rates
          - Table 3: Defence expenditure as share of GDP
            * Section 1 (rows 5-37): Share of real GDP (%)
            * Section 2 (rows 41-73): Annual real change (%)

        Returns:
            List of NATOSpendingRecord, one per country per year.
        """
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

        # --- Parse year headers ---
        ws2 = wb["Table 2"]
        years = self._parse_year_headers(ws2, header_row=4)

        # --- Table 2: spending USD ---
        spending_current = self._parse_section(ws2, _COUNTRY_ROWS_SECTION1, years)
        spending_current_agg = self._parse_section(
            ws2, _AGGREGATE_ROWS_SECTION1, years
        )
        spending_constant = self._parse_section(ws2, _COUNTRY_ROWS_SECTION2, years)
        spending_constant_agg = self._parse_section(
            ws2, _AGGREGATE_ROWS_SECTION2, years
        )

        # --- Table 3: % GDP and annual real change ---
        ws3 = wb["Table 3"]
        pct_gdp = self._parse_section(ws3, _COUNTRY_ROWS_SECTION1, years)
        pct_gdp_agg = self._parse_section(ws3, _AGGREGATE_ROWS_SECTION1, years)
        real_change = self._parse_section(ws3, _COUNTRY_ROWS_SECTION2, years)
        real_change_agg = self._parse_section(ws3, _AGGREGATE_ROWS_SECTION2, years)

        # --- Merge into records ---
        records: list[NATOSpendingRecord] = []

        # Individual countries
        all_countries = set(spending_current.keys()) | set(pct_gdp.keys())
        for country in sorted(all_countries):
            for year_col, (year, is_est) in enumerate(years):
                records.append(NATOSpendingRecord(
                    country=country,
                    year=year,
                    spending_current_usd_millions=self._get_val(
                        spending_current, country, year_col
                    ),
                    spending_constant_usd_millions=self._get_val(
                        spending_constant, country, year_col
                    ),
                    pct_gdp=self._get_val(pct_gdp, country, year_col),
                    annual_real_change_pct=self._get_val(
                        real_change, country, year_col
                    ),
                    is_estimate=is_est,
                    is_aggregate=False,
                ))

        # Aggregate rows
        all_agg = set(spending_current_agg.keys()) | set(pct_gdp_agg.keys())
        for agg_name in sorted(all_agg):
            for year_col, (year, is_est) in enumerate(years):
                records.append(NATOSpendingRecord(
                    country=agg_name,
                    year=year,
                    spending_current_usd_millions=self._get_val(
                        spending_current_agg, agg_name, year_col
                    ),
                    spending_constant_usd_millions=self._get_val(
                        spending_constant_agg, agg_name, year_col
                    ),
                    pct_gdp=self._get_val(pct_gdp_agg, agg_name, year_col),
                    annual_real_change_pct=self._get_val(
                        real_change_agg, agg_name, year_col
                    ),
                    is_estimate=is_est,
                    is_aggregate=True,
                ))

        logger.info(
            "Parsed %d NATO spending records (%d countries, %d years)",
            len(records),
            len(all_countries),
            len(years),
        )
        return records

    async def fetch_spending_data(
        self, url: str = NATO_EXCEL_URL
    ) -> list[NATOSpendingRecord]:
        """Download, parse, and return NATO spending data. Uses in-memory cache.

        Returns:
            List of NATOSpendingRecord for all countries and years.
        """
        if self._cache is not None:
            age = time.time() - self._cache.fetched_at
            if age < self.cache_ttl:
                logger.info(
                    "Returning cached NATO data (%d records, %.0fs old)",
                    len(self._cache.records),
                    age,
                )
                return self._cache.records

        excel_bytes = await self.fetch_excel(url)
        records = self.parse_excel(excel_bytes)

        self._cache = _CacheEntry(records=records, fetched_at=time.time())
        return records

    # ------------------------------------------------------------------ #
    # Internal helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _parse_year_headers(
        ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int
    ) -> list[tuple[int, bool]]:
        """Read year values from the header row.

        Returns list of (year, is_estimate) tuples, one per data column.
        Years ending with 'e' (e.g. '2024e') are flagged as estimates.
        """
        years: list[tuple[int, bool]] = []
        for col in range(_YEAR_COL_START, _YEAR_COL_END + 1):
            raw = ws.cell(row=header_row, column=col).value
            if raw is None:
                continue
            raw_str = str(raw).strip()
            is_estimate = raw_str.endswith("e")
            year_str = raw_str.rstrip("e")
            try:
                years.append((int(year_str), is_estimate))
            except ValueError:
                logger.warning("Unparseable year header in col %d: %r", col, raw)
        return years

    @staticmethod
    def _clean_country_name(raw: str) -> str:
        """Normalize a country name from the spreadsheet.

        Strips trailing asterisks, whitespace, and footnote markers.
        """
        name = str(raw).strip()
        # Remove trailing asterisks and other footnote markers
        name = re.sub(r"[*]+$", "", name).strip()
        return name

    def _parse_section(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row_range: range,
        years: list[tuple[int, bool]],
    ) -> dict[str, list[float | None]]:
        """Parse a block of country rows into {country: [values_per_year]}.

        Args:
            ws: The worksheet to read from.
            row_range: Range of 1-indexed row numbers to read.
            years: List of (year, is_estimate) for column alignment.

        Returns:
            Dict mapping cleaned country name to a list of float values
            (one per year column, None if missing).
        """
        data: dict[str, list[float | None]] = {}

        for row_idx in row_range:
            raw_name = ws.cell(row=row_idx, column=1).value
            if raw_name is None:
                continue

            country = self._clean_country_name(raw_name)
            if not country:
                continue

            values: list[float | None] = []
            for col_offset in range(len(years)):
                col = _YEAR_COL_START + col_offset
                cell_val = ws.cell(row=row_idx, column=col).value
                values.append(self._safe_float(cell_val))

            data[country] = values

        return data

    @staticmethod
    def _get_val(
        data: dict[str, list[float | None]],
        country: str,
        year_idx: int,
    ) -> float | None:
        """Safely retrieve a value from parsed section data."""
        vals = data.get(country)
        if vals is None or year_idx >= len(vals):
            return None
        return vals[year_idx]

    @staticmethod
    def _safe_float(value: object) -> float | None:
        """Convert a cell value to float, returning None on failure."""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
